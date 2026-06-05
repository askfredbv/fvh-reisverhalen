import csv
from collections import Counter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DL="C:/claude/fvh.com/downloads"; EX="C:/claude/fvh.com/exports"
rev=list(csv.DictReader(open(DL+"/bhag-reviews-flat.csv",encoding="utf-8-sig")))
trips=list(csv.DictReader(open(DL+"/trips-from-timeline.csv",encoding="utf-8")))

plek=Counter(); txt=Counter()
for r in rev:
    if str(r["reis_idx"]).strip().isdigit():
        i=int(r["reis_idx"]); plek[i]+=1
        if r["tekst"].strip(): txt[i]+=1

wb=Workbook()
ws=wb.active; ws.title="Reizen"
ws.append(["#","Van","Tot","Dagen","Regio","Max km","Plekken","Met tekst"])
for i,t in enumerate(trips):
    ws.append([i,t["van"],t["tot"],int(t["dagen"]),t["regio"].strip('"'),int(t["max_km"]),plek.get(i,0),txt.get(i,0)])
# totaalrij
ws.append(["","","TOTAAL",sum(int(t["dagen"]) for t in trips),"","",sum(plek.values()),sum(txt.values())])

ws2=wb.create_sheet("Plaatsen")
ws2.append(["Geschreven","Reis","Naam","Land","Adres","Lat","Lng","Sterren","Review-tekst","Maps URL","reis_idx"])
for row in rev:
    ws2.append([row["geschreven"],row["reis"],row["naam"],row["land"],row["adres"],
        float(row["lat"]) if row["lat"] else None, float(row["lng"]) if row["lng"] else None,
        int(row["sterren"]) if str(row["sterren"]).isdigit() else row["sterren"],
        row["tekst"],row["maps_url"],
        int(row["reis_idx"]) if str(row["reis_idx"]).strip().isdigit() else None])

def hdr(w,n):
    for c in range(1,n+1):
        cell=w.cell(1,c); cell.font=Font(name="Arial",bold=True,color="FFFFFF")
        cell.fill=PatternFill("solid",start_color="1F4E78"); cell.alignment=Alignment(vertical="center")
    w.freeze_panes="A2"; w.auto_filter.ref=w.dimensions
    for r in w.iter_rows(min_row=2):
        for cell in r: cell.font=Font(name="Arial")
hdr(ws,8); hdr(ws2,11)
# totaalrij bold
last=ws.max_row
for c in range(1,9): ws.cell(last,c).font=Font(name="Arial",bold=True)
for i,wd in enumerate([4,12,12,7,16,8,9,9],1): ws.column_dimensions[get_column_letter(i)].width=wd
for i,wd in enumerate([12,28,30,6,42,9,9,8,70,38,8],1): ws2.column_dimensions[get_column_letter(i)].width=wd
out=EX+"/BHAG-reizen-master.xlsx"; wb.save(out)

# verify read-back
wb2=load_workbook(out)
print("OK — bladen:",wb2.sheetnames,"| Reizen rijen:",wb2['Reizen'].max_row,"| Plaatsen rijen:",wb2['Plaatsen'].max_row)
print("saved",out)
